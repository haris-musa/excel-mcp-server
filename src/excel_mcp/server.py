import csv
import logging
import os
from typing import Any

from mcp.server.fastmcp import FastMCP

from excel_mcp.chart import create_chart_in_sheet as create_chart_impl
from excel_mcp.data import write_data

# Import exceptions
from excel_mcp.exceptions import (
    CalculationError,
    ChartError,
    DataError,
    FormattingError,
    PivotError,
    SheetError,
    ValidationError,
    WorkbookError,
)
from excel_mcp.pivot import create_pivot_table as create_pivot_table_impl
from excel_mcp.sheet import (
    copy_sheet,
    delete_sheet,
    merge_range,
    rename_sheet,
    unmerge_range,
)

# Import from excel_mcp package with consistent _impl suffixes
from excel_mcp.validation import (
    validate_formula_in_cell_operation as validate_formula_impl,
)
from excel_mcp.validation import (
    validate_range_in_sheet_operation as validate_range_impl,
)
from excel_mcp.workbook import get_workbook_info

# Get project root directory path for log file path.
# When using the stdio transmission method,
# relative paths may cause log files to fail to create
# due to the client's running location and permission issues,
# resulting in the program not being able to run.
# Thus using os.path.join(ROOT_DIR, "excel-mcp.log") instead.

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LOG_FILE = os.path.join(ROOT_DIR, "excel-mcp.log")

# Initialize EXCEL_FILES_PATH variable without assigning a value
EXCEL_FILES_PATH = None

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        # Referring to https://github.com/modelcontextprotocol/python-sdk/issues/409#issuecomment-2816831318
        # The stdio mode server MUST NOT write anything to its stdout that is not a valid MCP message.
        logging.FileHandler(LOG_FILE)
    ],
)
logger = logging.getLogger("excel-mcp")
# Initialize FastMCP server
mcp = FastMCP(
    "excel-mcp",
    version="0.1.3",
    description="Excel MCP Server for manipulating Excel files",
    dependencies=["openpyxl>=3.1.2"],
    env_vars={
        "EXCEL_FILES_PATH": {
            "description": "Path to Excel files directory",
            "required": False,
            "default": EXCEL_FILES_PATH,
        }
    },
)


def get_excel_path(filename: str) -> str:
    """Get full path to Excel file.

    Args:
        filename: Name of Excel file

    Returns:
        Full path to Excel file
    """
    # If filename is already an absolute path, return it
    if os.path.isabs(filename):
        return filename

    # Check if in SSE mode (EXCEL_FILES_PATH is not None)
    if EXCEL_FILES_PATH is None:
        # Must use absolute path
        raise ValueError(
            f"Invalid filename: {filename}, must be an absolute path when not in SSE mode"
        )

    # In SSE mode, if it's a relative path, resolve it based on EXCEL_FILES_PATH
    return os.path.join(EXCEL_FILES_PATH, filename)


@mcp.tool()
def apply_formula(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """
    Apply Excel formula to cell.
    Excel formula will write to cell with verification.
    """
    try:
        full_path = get_excel_path(filepath)
        # First validate the formula
        validation = validate_formula_impl(full_path, sheet_name, cell, formula)
        if isinstance(validation, dict) and "error" in validation:
            return f"Error: {validation['error']}"

        # If valid, apply the formula
        from excel_mcp.calculations import apply_formula as apply_formula_impl

        result = apply_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error applying formula: {e}")
        raise


@mcp.tool()
def validate_formula_syntax(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """Validate Excel formula syntax without applying it."""
    try:
        full_path = get_excel_path(filepath)
        result = validate_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating formula: {e}")
        raise


@mcp.tool()
def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: int = None,
    font_color: str = None,
    bg_color: str = None,
    border_style: str = None,
    border_color: str = None,
    number_format: str = None,
    alignment: str = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: dict[str, Any] = None,
    conditional_format: dict[str, Any] = None,
) -> str:
    """Apply formatting to a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.formatting import format_range as format_range_func

        _ = format_range_func(
            filepath=full_path,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,
            font_color=font_color,
            bg_color=bg_color,
            border_style=border_style,
            border_color=border_color,
            number_format=number_format,
            alignment=alignment,
            wrap_text=wrap_text,
            merge_cells=merge_cells,
            protection=protection,
            conditional_format=conditional_format,
        )
        return "Range formatted successfully"
    except (ValidationError, FormattingError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error formatting range: {e}")
        raise


@mcp.tool()
def read_data_from_excel(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: str = None,
    preview_only: bool = False,
) -> str:
    """
    Read data from Excel worksheet.

    Returns:
    Data from Excel worksheet as json string. list of lists or empty list if no data found. sublists are assumed to be rows.
    """
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.data import read_excel_range

        result = read_excel_range(
            full_path, sheet_name, start_cell, end_cell, preview_only
        )
        if not result:
            return "No data found in specified range"
        # Convert the list of dicts to a formatted string
        data_str = "\n".join([str(row) for row in result])
        return data_str
    except Exception as e:
        logger.error(f"Error reading data: {e}")
        raise


@mcp.tool()
def write_data_to_excel(
    filepath: str,
    sheet_name: str,
    data: list[list],
    start_cell: str = "A1",
    auto_detect_types: bool = True,
) -> str:
    """
    Write data to Excel worksheet with automatic data type detection.

    PARAMETERS:
    filepath: Path to Excel file
    sheet_name: Name of worksheet to write to
    data: List of lists containing data to write to the worksheet, sublists are assumed to be rows
    start_cell: Cell to start writing to, default is "A1"
    auto_detect_types: Whether to automatically detect and convert data types (numbers, dates, etc.)

    When auto_detect_types is True, the function will:
    - Convert string numbers to actual numbers (e.g., "123" -> 123)
    - Detect and format percentages (e.g., "50%" -> 0.5 with percentage format)
    - Detect and format currency (e.g., "$1,000" -> 1000 with currency format)
    - Detect and format dates (e.g., "2023-12-25" -> date with date format)
    - Convert boolean-like strings (e.g., "true" -> TRUE)
    """
    try:
        full_path = get_excel_path(filepath)
        result = write_data(full_path, sheet_name, data, start_cell, auto_detect_types)
        return result["message"]
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error writing data: {e}")
        raise


@mcp.tool()
def write_csv_to_excel(
    filepath: str,
    sheet_name: str,
    csv_path: str,
    start_cell: str = "A1",
) -> str:
    """
    Write CSV file data to Excel worksheet.

    PARAMETERS:
    filepath: Path to Excel file
    sheet_name: Name of worksheet to write to
    csv_path: Path to CSV file to read from
    start_cell: Cell to start writing to, default is "A1"
    """
    try:
        full_path = get_excel_path(filepath)

        # Read CSV data
        data = []
        with open(csv_path, "r", newline="", encoding="utf-8") as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                data.append(row)

        # Write data to Excel using the existing write_data function with type detection
        result = write_data(
            full_path, sheet_name, data, start_cell, auto_detect_types=True
        )
        return result["message"]
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except FileNotFoundError:
        return f"Error: CSV file not found: {csv_path}"
    except Exception as e:
        logger.error(f"Error writing CSV to Excel: {e}")
        raise


@mcp.tool()
def create_workbook(filepath: str) -> str:
    """Create new Excel workbook."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_workbook as create_workbook_impl

        _ = create_workbook_impl(full_path)
        return f"Created workbook at {full_path}"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")
        raise


@mcp.tool()
def create_worksheet(filepath: str, sheet_name: str) -> str:
    """Create new worksheet in workbook."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.workbook import create_sheet as create_worksheet_impl

        result = create_worksheet_impl(full_path, sheet_name)
        return result["message"]
    except (ValidationError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating worksheet: {e}")
        raise


@mcp.tool()
def create_chart(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
) -> str:
    """Create chart in worksheet."""
    try:
        full_path = get_excel_path(filepath)
        result = create_chart_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            chart_type=chart_type,
            target_cell=target_cell,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis,
        )
        return result["message"]
    except (ValidationError, ChartError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        raise


@mcp.tool()
def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: list[str],
    values: list[str],
    columns: list[str] = None,
    agg_func: str = "mean",
) -> str:
    """Create pivot table in worksheet."""
    try:
        full_path = get_excel_path(filepath)
        result = create_pivot_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            rows=rows,
            values=values,
            columns=columns or [],
            agg_func=agg_func,
        )
        return result["message"]
    except (ValidationError, PivotError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        raise


@mcp.tool()
def copy_worksheet(filepath: str, source_sheet: str, target_sheet: str) -> str:
    """Copy worksheet within workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = copy_sheet(full_path, source_sheet, target_sheet)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying worksheet: {e}")
        raise


@mcp.tool()
def delete_worksheet(filepath: str, sheet_name: str) -> str:
    """Delete worksheet from workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = delete_sheet(full_path, sheet_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting worksheet: {e}")
        raise


@mcp.tool()
def rename_worksheet(filepath: str, old_name: str, new_name: str) -> str:
    """Rename worksheet in workbook."""
    try:
        full_path = get_excel_path(filepath)
        result = rename_sheet(full_path, old_name, new_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error renaming worksheet: {e}")
        raise


@mcp.tool()
def get_workbook_metadata(filepath: str, include_ranges: bool = False) -> str:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        full_path = get_excel_path(filepath)
        result = get_workbook_info(full_path, include_ranges=include_ranges)
        return str(result)
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting workbook metadata: {e}")
        raise


@mcp.tool()
def merge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Merge a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        result = merge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error merging cells: {e}")
        raise


@mcp.tool()
def unmerge_cells(
    filepath: str, sheet_name: str, start_cell: str, end_cell: str
) -> str:
    """Unmerge a range of cells."""
    try:
        full_path = get_excel_path(filepath)
        result = unmerge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error unmerging cells: {e}")
        raise


@mcp.tool()
def copy_range(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: str = None,
) -> str:
    """Copy a range of cells to another location."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import copy_range_operation

        result = copy_range_operation(
            full_path, sheet_name, source_start, source_end, target_start, target_sheet
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying range: {e}")
        raise


@mcp.tool()
def delete_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    shift_direction: str = "up",
) -> str:
    """Delete a range of cells and shift remaining cells."""
    try:
        full_path = get_excel_path(filepath)
        from excel_mcp.sheet import delete_range_operation

        result = delete_range_operation(
            full_path, sheet_name, start_cell, end_cell, shift_direction
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting range: {e}")
        raise


@mcp.tool()
def validate_excel_range(
    filepath: str, sheet_name: str, start_cell: str, end_cell: str = None
) -> str:
    """Validate if a range exists and is properly formatted."""
    try:
        full_path = get_excel_path(filepath)
        range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        result = validate_range_impl(full_path, sheet_name, range_str)
        return result["message"]
    except ValidationError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating range: {e}")
        raise


@mcp.tool()
def auto_format_data_range(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: str = None,
) -> str:
    """
    Automatically detect and apply proper formatting to data in a range.

    This function analyzes existing cell values and applies appropriate number formatting:
    - Detects numbers and applies numeric formatting
    - Detects percentages and applies percentage formatting
    - Detects currency and applies currency formatting
    - Detects dates and applies date formatting

    PARAMETERS:
    filepath: Path to Excel file
    sheet_name: Name of worksheet
    start_cell: Starting cell reference (default: A1)
    end_cell: Ending cell reference (optional, if not provided will format all data in sheet)
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        from excel_mcp.cell_utils import parse_cell_range
        from excel_mcp.data import _infer_and_convert_data_type

        wb = load_workbook(full_path)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"

        ws = wb[sheet_name]

        # Determine range to format
        if end_cell:
            try:
                start_row, start_col, end_row, end_col = parse_cell_range(
                    start_cell, end_cell
                )
            except ValueError as e:
                return f"Error: Invalid cell range: {str(e)}"
        else:
            # Use all data in the sheet
            start_row, start_col = 1, 1
            end_row, end_col = ws.max_row, ws.max_column

        formatted_count = 0

        # Process each cell in the range
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    # Analyze the current value and determine proper formatting
                    converted_value, number_format = _infer_and_convert_data_type(
                        cell.value
                    )

                    # Apply formatting if detected
                    if number_format and converted_value is not None:
                        cell.value = converted_value
                        cell.number_format = number_format
                        formatted_count += 1

        wb.save(full_path)
        wb.close()

        range_desc = (
            f"{start_cell}:{end_cell}"
            if end_cell
            else f"{start_cell} to {get_column_letter(end_col)}{end_row}"
        )
        return f"Applied automatic formatting to {formatted_count} cells in range {range_desc}"

    except Exception as e:
        logger.error(f"Error auto-formatting data: {e}")
        return f"Error: {str(e)}"


async def run_sse():
    """Run Excel MCP server in SSE mode."""
    # Assign value to EXCEL_FILES_PATH in SSE mode
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    # Create directory if it doesn't exist
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)

    try:
        logger.info(
            f"Starting Excel MCP server with SSE transport (files directory: {EXCEL_FILES_PATH})"
        )
        await mcp.run_sse_async()
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
        await mcp.shutdown()
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")


def run_stdio():
    """Run Excel MCP server in stdio mode."""
    # No need to assign EXCEL_FILES_PATH in stdio mode

    try:
        logger.info("Starting Excel MCP server with stdio transport")
        mcp.run(transport="stdio")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")
