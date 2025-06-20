import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .cell_utils import parse_cell_range
from .cell_validation import get_data_validation_for_cell
from .exceptions import DataError

logger = logging.getLogger(__name__)


def _infer_and_convert_data_type(value: Any) -> tuple[Any, str | None]:
    """
    Infer the data type of a value and convert it appropriately for Excel.

    Returns:
        tuple: (converted_value, number_format)
    """
    if value is None or value == "":
        return None, None

    # Convert to string for analysis
    str_value = str(value).strip()

    if not str_value:
        return None, None

    # Try to detect and convert numeric values
    try:
        # Check for percentage
        if str_value.endswith("%"):
            numeric_part = str_value[:-1].replace(",", "")
            if numeric_part.replace(".", "").replace("-", "").isdigit():
                return float(numeric_part) / 100, "0.00%"

        # Check for currency (basic detection)
        currency_pattern = r"^[\$€£¥]?[\d,]+\.?\d*$"
        if re.match(currency_pattern, str_value):
            numeric_part = re.sub(r"[\$€£¥,]", "", str_value)
            if "." in numeric_part:
                return float(numeric_part), '"$"#,##0.00'
            else:
                return int(numeric_part), '"$"#,##0'

        # Check for regular numbers (with comma separators)
        number_pattern = r"^-?[\d,]+\.?\d*$"
        if re.match(number_pattern, str_value):
            numeric_part = str_value.replace(",", "")
            if "." in numeric_part:
                return float(numeric_part), "0.00"
            else:
                return int(numeric_part), "0"

        # Try direct conversion for scientific notation, etc.
        if "." in str_value or "e" in str_value.lower():
            return float(str_value), "0.00"
        else:
            return int(str_value), "0"

    except (ValueError, TypeError):
        pass

    # Try to detect dates
    try:
        # Common date formats
        date_formats = [
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%m/%d/%Y %H:%M",
            "%d/%m/%Y %H:%M",
        ]

        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(str_value, fmt)
                if "%H" in fmt:  # Has time component
                    return parsed_date, "mm/dd/yyyy h:mm"
                else:
                    return parsed_date.date(), "mm/dd/yyyy"
            except ValueError:
                continue

    except (ValueError, TypeError):
        pass

    # Check for boolean values
    if str_value.lower() in ["true", "false", "yes", "no", "1", "0"]:
        if str_value.lower() in ["true", "yes", "1"]:
            return True, None
        else:
            return False, None

    # Return as text if no conversion possible
    return str_value, None


def read_excel_range(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: str | None = None,
    preview_only: bool = False,
) -> list[dict[str, Any]]:
    """Read data from Excel range with optional preview mode"""
    try:
        wb = load_workbook(filepath, read_only=False)

        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        # Parse start cell
        if ":" in start_cell:
            start_cell, end_cell = start_cell.split(":")

        # Get start coordinates
        try:
            start_coords = parse_cell_range(f"{start_cell}:{start_cell}")
            if not start_coords or not all(
                coord is not None for coord in start_coords[:2]
            ):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Determine end coordinates
        if end_cell:
            try:
                end_coords = parse_cell_range(f"{end_cell}:{end_cell}")
                if not end_coords or not all(
                    coord is not None for coord in end_coords[:2]
                ):
                    raise DataError(f"Invalid end cell reference: {end_cell}")
                end_row, end_col = end_coords[0], end_coords[1]
            except ValueError as e:
                raise DataError(f"Invalid end cell format: {str(e)}")
        else:
            # If no end_cell, use the full data range of the sheet
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                # Handle empty sheet
                end_row, end_col = start_row, start_col
            else:
                # Use the sheet's own boundaries
                start_row, start_col = ws.min_row, ws.min_column
                end_row, end_col = ws.max_row, ws.max_column

        # Validate range bounds
        if start_row > ws.max_row or start_col > ws.max_column:
            # This case can happen if start_cell is outside the used area on a sheet with data
            # or on a completely empty sheet.
            logger.warning(
                f"Start cell {start_cell} is outside the sheet's data boundary "
                f"({get_column_letter(ws.min_column)}{ws.min_row}:{get_column_letter(ws.max_column)}{ws.max_row}). "
                f"No data will be read."
            )
            return []

        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            if any(v is not None for v in row_data):
                data.append(row_data)

        wb.close()
        return data
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to read Excel range: {e}")
        raise DataError(str(e))


def _write_data_to_worksheet(
    worksheet: Worksheet,
    data: list[list],
    start_cell: str = "A1",
    auto_detect_types: bool = True,
) -> None:
    """Write data to worksheet with intelligent header handling and data type detection"""
    try:
        if not data:
            raise DataError("No data provided to write")

        try:
            start_coords = parse_cell_range(start_cell)
            if not start_coords or not all(x is not None for x in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Write data with type detection and formatting
        for i, row in enumerate(data):
            for j, val in enumerate(row):
                cell = worksheet.cell(row=start_row + i, column=start_col + j)

                if auto_detect_types:
                    converted_value, number_format = _infer_and_convert_data_type(val)
                    cell.value = converted_value

                    # Apply number formatting if detected
                    if number_format:
                        cell.number_format = number_format
                else:
                    cell.value = val

    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write worksheet data: {e}")
        raise DataError(str(e))


def write_data(
    filepath: str,
    sheet_name: str | None,
    data: list[list] | None,
    start_cell: str = "A1",
    auto_detect_types: bool = True,
) -> dict[str, str]:
    """Write data to Excel sheet with workbook handling and enhanced type detection

    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        data: Data to write
        start_cell: Starting cell reference
        auto_detect_types: Whether to automatically detect and convert data types

    Headers are handled intelligently based on context.
    """
    try:
        if not data:
            raise DataError("No data provided to write")

        wb = load_workbook(filepath)

        # If no sheet specified, use active sheet
        if not sheet_name:
            sheet_name = wb.active.title
        elif sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]

        # Validate start cell
        try:
            start_coords = parse_cell_range(start_cell)
            if not start_coords or not all(
                coord is not None for coord in start_coords[:2]
            ):
                raise DataError(f"Invalid start cell reference: {start_cell}")
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        if len(data) > 0:
            _write_data_to_worksheet(ws, data, start_cell, auto_detect_types)

        wb.save(filepath)
        wb.close()

        return {
            "message": f"Data written to {sheet_name} with type detection",
            "active_sheet": sheet_name,
        }
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write data: {e}")
        raise DataError(str(e))


def read_excel_range_with_metadata(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: str | None = None,
    include_validation: bool = True,
) -> dict[str, Any]:
    """Read data from Excel range with cell metadata including validation rules.

    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell address
        end_cell: Ending cell address (optional)
        include_validation: Whether to include validation metadata

    Returns:
        Dictionary containing structured cell data with metadata
    """
    try:
        wb = load_workbook(filepath, read_only=False)

        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        # Parse start cell
        if ":" in start_cell:
            start_cell, end_cell = start_cell.split(":")

        # Get start coordinates
        try:
            start_coords = parse_cell_range(f"{start_cell}:{start_cell}")
            if not start_coords or not all(
                coord is not None for coord in start_coords[:2]
            ):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Determine end coordinates
        if end_cell:
            try:
                end_coords = parse_cell_range(f"{end_cell}:{end_cell}")
                if not end_coords or not all(
                    coord is not None for coord in end_coords[:2]
                ):
                    raise DataError(f"Invalid end cell reference: {end_cell}")
                end_row, end_col = end_coords[0], end_coords[1]
            except ValueError as e:
                raise DataError(f"Invalid end cell format: {str(e)}")
        else:
            # If no end_cell, use the full data range of the sheet
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                # Handle empty sheet
                end_row, end_col = start_row, start_col
            else:
                # Use the sheet's own boundaries, but respect the provided start_cell
                end_row, end_col = ws.max_row, ws.max_column
                # If start_cell is 'A1' (default), we should find the true start
                if start_cell == "A1":
                    start_row, start_col = ws.min_row, ws.min_column

        # Validate range bounds
        if start_row > ws.max_row or start_col > ws.max_column:
            # This case can happen if start_cell is outside the used area on a sheet with data
            # or on a completely empty sheet.
            logger.warning(
                f"Start cell {start_cell} is outside the sheet's data boundary "
                f"({get_column_letter(ws.min_column)}{ws.min_row}:{get_column_letter(ws.max_column)}{ws.max_row}). "
                f"No data will be read."
            )
            return {"range": f"{start_cell}:", "sheet_name": sheet_name, "cells": []}

        # Build structured cell data
        range_str = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        range_data = {"range": range_str, "sheet_name": sheet_name, "cells": []}

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell_address = f"{get_column_letter(col)}{row}"

                cell_data = {
                    "address": cell_address,
                    "value": cell.value,
                    "row": row,
                    "column": col,
                }

                # Add validation metadata if requested
                if include_validation:
                    validation_info = get_data_validation_for_cell(ws, cell_address)
                    if validation_info:
                        cell_data["validation"] = validation_info
                    else:
                        cell_data["validation"] = {"has_validation": False}

                range_data["cells"].append(cell_data)

        wb.close()
        return range_data

    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to read Excel range with metadata: {e}")
        raise DataError(str(e))
